import sys
import os
import re
import json
import pytesseract
from configparser import ConfigParser

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QFileDialog, QLabel, QTableWidget, QTableWidgetItem, QMessageBox,
    QHeaderView, QDialog, QLineEdit, QDialogButtonBox,
    QGridLayout, QAbstractItemView, QMenu
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QColor, QPixmap, QIcon, QAction
from openpyxl import Workbook
from openpyxl.styles import Font as ExcelFont, PatternFill, Alignment, Side, Border
from openpyxl.styles.colors import Color

# --- Check for TextWordWrap attribute
WORD_WRAP_FLAG = 0
if hasattr(Qt, 'TextWordWrap'):
    WORD_WRAP_FLAG = Qt.TextWordWrap
elif hasattr(Qt.AlignmentFlag, 'TextWordWrap'):
    WORD_WRAP_FLAG = Qt.AlignmentFlag.TextWordWrap
else:
    print("Warning: TextWordWrap flag not found. Text wrapping may not work correctly.")


APP_STYLESHEET = """
QWidget { background-color: #333333; color: #C8CDD3; font-family: 'Segoe UI', 'Roboto', sans-serif; font-size: 12px; }
QLabel#title_label { color: #00BFFF; font-size: 28px; font-weight: bold; margin-bottom: 20px; }
QPushButton { background-color: #30363D; color: #FFFFFF; border: 1px solid #00BFFF; border-radius: 8px; padding: 8px 15px; font-weight: bold; text-transform: uppercase; }
QPushButton:hover { background-color: #3B424A; border: 1px solid #4AC8FF; }
QPushButton:pressed { background-color: #00BFFF; color: #21252B; }
QPushButton:disabled { background-color: #282C34; color: #636D83; border: 1px solid #282C34; }
QTableWidget { background-color: #21252B; alternate-background-color: #2A2F36; color: #C8CDD3; gridline-color: #30363D; border: 1px solid #00BFFF; selection-background-color: #00BFFF; selection-color: #21252B; border-radius: 8px; }
QTableWidget::item { padding: 4px; }
QHeaderView::section { background-color: #30363D; color: #00BFFF; padding: 6px; border: 1px solid #21252B; font-weight: bold; text-transform: uppercase; }
QHeaderView::section:horizontal { border-top: 0px solid #21252B; border-bottom: 1px solid #00BFFF; }
QMessageBox { background-color: #21252B; color: #C8CDD3; font-size: 12px; }
QMessageBox QPushButton { background-color: #00BFFF; color: #21252B; border: none; padding: 7px 12px; border-radius: 5px; font-weight: bold; }
QMessageBox QPushButton:hover { background-color: #4AC8FF; color: #21252B; }
QLineEdit { background-color: #30363D; color: #C8CDD3; border: 1px solid #00BFFF; border-radius: 5px; padding: 5px; }
QLabel { color: #C8CDD3; }
QCheckBox { color: #C8CDD3; spacing: 5px; }
QCheckBox::indicator { width: 16px; height: 16px; border-radius: 3px; border: 1px solid #00BFFF; background-color: #30363D; }
QCheckBox::indicator:checked { background-color: #00BFFF; border: 1px solid #4AC8FF; }
QCheckBox::indicator:hover { border: 1px solid #4AC8FF; }

/* Desain scrollbar yang lebih kecil dan estetik */
QScrollBar:vertical {
    border: none;
    background: #21252B;
    width: 10px;
    margin: 0px 0px 0px 0px;
}
QScrollBar::handle:vertical {
    background: #4A4A4A;
    border: 1px solid #303030;
    min-height: 20px;
    border-radius: 5px;
}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
    border: none;
    background: none;
}
QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
    background: none;
}

QScrollBar:horizontal {
    border: none;
    background: #21252B;
    height: 10px;
    margin: 0px 0px 0px 0px;
}
QScrollBar::handle:horizontal {
    background: #4A4A4A;
    border: 1px solid #303030;
    min-width: 20px;
    border-radius: 5px;
}
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
    border: none;
    background: none;
}
QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {
    background: none;
}
"""

def get_resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# --- KONSTANTA PENTING ---
EXTRACTION_KEYWORD_PRIMARY = "Diagnosa"
REGEX_EXTRACTION_PRIMARY = re.compile(
    rf"{EXTRACTION_KEYWORD_PRIMARY}(?:\s*Utama|\s*\d\.?)?\s*[:;,-]?\s*(.*?)(?={EXTRACTION_KEYWORD_PRIMARY}(?:\s*Sekunder|\s*\d\.?)?|Validasi hasil|\Z)",
    re.IGNORECASE | re.DOTALL
)
REGEX_SPLIT_DIAGNOSA = re.compile(r"^(\s*([a-zA-Z]\d{2}(?:\.\d{1,2})?))(?:\s*-\s*|\s*)(.*)", re.IGNORECASE)

# --- Kelas Worker untuk Memproses PDF (Berjalan di Thread Terpisah) ---
class PdfProcessingWorker(QThread):
    finished = pyqtSignal(str, dict, str, dict, bool)
    keyword_found_signal = pyqtSignal(str, str, int)

    def __init__(self, pdf_path, texts_to_find_tuples, dpi, validation_rules, code_text):
        super().__init__()
        self.pdf_path = pdf_path
        self.texts_to_find_tuples = texts_to_find_tuples
        self.dpi = dpi
        self.validation_rules = validation_rules
        self.code_text = code_text

    def _get_page_text(self, page, page_num, dpi):
        page_text_from_pdf = page.get_text()
        if page_text_from_pdf.strip():
            return page_text_from_pdf
        
        if not pytesseract:
            return ""

        try:
            zoom = dpi / 72.0
            matrix = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=matrix)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            osd_data = pytesseract.image_to_osd(img)
            rotation_angle_match = re.search(r"Rotate: (\d+)", osd_data)
            if rotation_angle_match:
                angle = int(rotation_angle_match.group(1))
                if angle != 0:
                    img = img.rotate(-angle, expand=True)
            config_ocr = '--psm 3 -l eng+ind'
            page_text_from_ocr = pytesseract.image_to_string(img, config=config_ocr)
            return page_text_from_ocr
        except pytesseract.TesseractNotFoundError:
            raise pytesseract.TesseractNotFoundError("Mesin Tesseract OCR tidak ditemukan.")
        except Exception as e:
            if "Too few characters" in str(e) or "Error during processing." in str(e) or "Invalid resolution" in str(e):
                return ""
            else:
                raise Exception(f"Error serius saat OCR di halaman {page_num + 1}: {e}")

    def run(self):
        grouped_search_texts = {}
        for search_text, display_text in self.texts_to_find_tuples:
            grouped_search_texts.setdefault(display_text, []).append(search_text)

        results = {display_text: False for display_text in grouped_search_texts.keys()}
        found_page_numbers = {display_text: -1 for display_text in grouped_search_texts.keys()}

        error_message = ""
        document = None
        found_ringkasan_keyword = False
        
        try:
            document = fitz.open(self.pdf_path)
            
            start_page_for_keywords = 0
            
            for page_num in range(document.page_count):
                page_content = self._get_page_text(document.load_page(page_num), page_num, self.dpi)
                if re.search(r'\b(kriteria\s+discharge\s+planing|rm\s*29|permintaan\s+rawat\s+inap|discharge\s+planing)\b', page_content.lower(), re.IGNORECASE | re.DOTALL):
                    start_page_for_keywords = page_num
                    found_ringkasan_keyword = True
                    break
            
            if not found_ringkasan_keyword:
                self.finished.emit(self.pdf_path, {}, "Tidak Lulus: 'Permintaan Rawat Inap' tidak ditemukan.", {}, found_ringkasan_keyword)
                return

            for display_text, search_texts in grouped_search_texts.items():
                for page_num in range(start_page_for_keywords, document.page_count):
                    page_content = self._get_page_text(document.load_page(page_num), page_num, self.dpi)
                    found_on_page = False
                    for search_text in search_texts:
                        if search_text.lower() in page_content.lower():
                            results[display_text] = True
                            found_page_numbers[display_text] = page_num + 1
                            self.keyword_found_signal.emit(self.pdf_path, display_text, page_num + 1)
                            found_on_page = True
                            break
                    if found_on_page:
                        break

            if self.code_text != "Tidak Ditemukan":
                code_text_lower = self.code_text.lower()
                for diag_keyword_regex, must_have_keyword in self.validation_rules:
                    if re.search(diag_keyword_regex, code_text_lower):
                        found_validation_keyword_page = -1
                        for i in range(document.page_count - 1, -1, -1):
                            page_content = self._get_page_text(document.load_page(i), i, self.dpi)
                            if must_have_keyword.lower() in page_content.lower():
                                found_validation_keyword_page = i + 1
                                break
                        found_page_numbers[must_have_keyword.upper()] = found_validation_keyword_page

        except fitz.FileNotFoundError:
            error_message = "File PDF tidak ditemukan."
        except pytesseract.TesseractNotFoundError:
            error_message = "Mesin Tesseract OCR tidak ditemukan."
        except Exception as e:
            error_message = f"Terjadi kesalahan umum saat memproses PDF: {e}"
        finally:
            if document:
                document.close()
        
        self.finished.emit(self.pdf_path, results, error_message, found_page_numbers, found_ringkasan_keyword)

# --- Kelas Dialog untuk Mengatur Kata Kunci ---
class DraggableTableWidget(QTableWidget):
    rowsMoved = pyqtSignal(list)

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setDragDropMode(QAbstractItemView.DragDropMode.InternalMove)
        self.setDragEnabled(True)
        self.setAcceptDrops(True)
        self.setDropIndicatorShown(True)
        self.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self.on_context_menu)

    def on_context_menu(self, pos):
        item = self.itemAt(pos)
        if item is not None and item.column() == 1:
            menu = QMenu(self)
            copy_action = QAction("Salin Nama File", self)
            copy_action.triggered.connect(lambda: self.copy_cell_text(item))
            menu.addAction(copy_action)
            menu.exec(self.mapToGlobal(pos))
    
    def keyPressEvent(self, event):
        if event.modifiers() == Qt.KeyboardModifier.ControlModifier and event.key() == Qt.Key.Key_C:
            selected_items = self.selectedItems()
            if selected_items:
                item = next((item for item in selected_items if item.column() == 1), None)
                if item:
                    self.copy_cell_text(item)
                    return
        
        super().keyPressEvent(event)
    
    def copy_cell_text(self, item):
        if item:
            text_to_copy = item.text()
            QApplication.clipboard().setText(text_to_copy)
            QMessageBox.information(self, "Salin Berhasil", f"Nama file '{text_to_copy}' berhasil disalin ke clipboard.")

    def dropEvent(self, event):
        if event.source() == self and (event.dropAction() == Qt.DropAction.MoveAction or event.proposedAction() == Qt.DropAction.MoveAction):
            selected_rows_indices = sorted([index.row() for index in self.selectedIndexes()])
            if not selected_rows_indices:
                event.ignore()
                return
            moved_items_data = []
            for r_idx in selected_rows_indices:
                row_data = [self.item(r_idx, col).text() for col in range(self.columnCount())]
                moved_items_data.append(row_data)
            drop_row = self.indexAt(event.position().toPoint()).row()
            if drop_row == -1: drop_row = self.rowCount()
            self.blockSignals(True)
            for r_idx in sorted(selected_rows_indices, reverse=True): self.removeRow(r_idx)
            for i, item_data in enumerate(moved_items_data):
                insert_at = drop_row + i
                self.insertRow(insert_at)
                for col, value in enumerate(item_data):
                    self.setItem(insert_at, col, QTableWidgetItem(value))
            self.blockSignals(False)
            new_order = [tuple(self.item(r, c).text() for c in range(self.columnCount())) for r in range(self.rowCount())]
            self.rowsMoved.emit(new_order)
            event.acceptProposedAction()
        else: super().dropEvent(event)

class KeywordManagerDialog(QDialog):
    def __init__(self, current_keywords_tuples, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Atur Kata Kunci Pencarian")
        self.setGeometry(200, 200, 600, 450)
        self.current_keywords = list(current_keywords_tuples)
        self.layout = QVBoxLayout()
        self.keyword_table = DraggableTableWidget()
        self.keyword_table.setColumnCount(2)
        self.keyword_table.setHorizontalHeaderLabels(["Teks Pencarian (Internal)", "Teks Tampilan (GUI)"])
        self.keyword_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.layout.addWidget(self.keyword_table)
        self.keyword_table.rowsMoved.connect(self._update_keywords_from_table_order)
        self._populate_table()
        input_grid_layout = QGridLayout()
        input_grid_layout.addWidget(QLabel("Teks Pencarian:"), 0, 0)
        self.search_input_field = QLineEdit()
        self.search_input_field.setPlaceholderText("Teks yang akan dicari di PDF (misal: 'Prosedur Non-Bedah')")
        input_grid_layout.addWidget(self.search_input_field, 0, 1)
        input_grid_layout.addWidget(QLabel("Teks Tampilan:"), 1, 0)
        self.display_input_field = QLineEdit()
        self.display_input_field.setPlaceholderText("Teks yang akan ditampilkan di tabel utama (misal: 'Non-Bedah')")
        input_grid_layout.addWidget(self.display_input_field, 1, 1)
        self.layout.addLayout(input_grid_layout)
        button_row = QHBoxLayout()

        self.add_button = QPushButton("Tambah")
        self.add_button.clicked.connect(self.add_keyword)
        
        self.edit_button = QPushButton("Edit Terpilih")
        self.edit_button.clicked.connect(self.edit_selected_keyword)
        self.edit_button.setEnabled(False)

        self.remove_button = QPushButton("Hapus Terpilih")
        self.remove_button.clicked.connect(self.remove_keyword)
        self.remove_button.setEnabled(False)

        button_row.addWidget(self.add_button)
        button_row.addWidget(self.edit_button)
        button_row.addWidget(self.remove_button)
        self.layout.addLayout(button_row)
        self.button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        self.layout.addWidget(self.button_box)
        self.setLayout(self.layout)
        self.keyword_table.itemSelectionChanged.connect(self.update_buttons_state)
        self.keyword_table.itemDoubleClicked.connect(self.edit_selected_keyword)
        if parent: self.setStyleSheet(parent.styleSheet())

    def _populate_table(self):
        self.keyword_table.setRowCount(0)
        for search_text, display_text in self.current_keywords:
            row_position = self.keyword_table.rowCount()
            self.keyword_table.insertRow(row_position)
            self.keyword_table.setItem(row_position, 0, QTableWidgetItem(search_text))
            self.keyword_table.setItem(row_position, 1, QTableWidgetItem(display_text))

    def _update_keywords_from_table_order(self, new_order_list):
        self.current_keywords = new_order_list

    def update_buttons_state(self):
        has_selection = bool(self.keyword_table.selectedItems())
        self.edit_button.setEnabled(has_selection)
        self.remove_button.setEnabled(has_selection)

    def add_keyword(self):
        search_text = self.search_input_field.text().strip()
        display_text = self.display_input_field.text().strip()
        if not search_text or not display_text:
            QMessageBox.warning(self, "Input Kosong", "Teks Pencarian dan Teks Tampilan tidak boleh kosong.")
            return
        new_keyword_tuple = (search_text, display_text)
        if any(kw[0].lower() == search_text.lower() for kw in self.current_keywords):
            QMessageBox.warning(self, "Duplikat", "Teks Pencarian sudah ada dalam daftar.")
            return
        self.current_keywords.append(new_keyword_tuple)
        self._populate_table()
        self.search_input_field.clear(); self.display_input_field.clear(); self.search_input_field.setFocus()
    
    def edit_selected_keyword(self):
        selected_items = self.keyword_table.selectedItems()
        if selected_items:
            row = selected_items[0].row()
            self.search_input_field.setText(self.keyword_table.item(row, 0).text())
            self.display_input_field.setText(self.keyword_table.item(row, 1).text())
            self.remove_keyword(skip_msg=True)
            self.search_input_field.setFocus()

    def remove_keyword(self, skip_msg=False):
        selected_rows = sorted(list(set(index.row() for index in self.keyword_table.selectedIndexes())), reverse=True)
        if selected_rows:
            for row_to_remove in selected_rows:
                if 0 <= row_to_remove < len(self.current_keywords): del self.current_keywords[row_to_remove]
            self._populate_table()
            if not skip_msg: self.search_input_field.clear(); self.display_input_field.clear()
            self.update_buttons_state()
        elif not skip_msg: QMessageBox.warning(self, "Tidak Ada Pilihan", "Pilih kata kunci yang ingin dihapus.")
    
    def get_updated_keywords(self): return self.current_keywords


# --- Kelas Dialog untuk Mengatur Aturan Validasi ---
class ValidationRuleManagerDialog(QDialog):
    def __init__(self, current_rules, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Atur Aturan Validasi")
        self.setGeometry(200, 200, 700, 450)
        self.current_rules = list(current_rules)
        self.layout = QVBoxLayout()
        self.rules_table = DraggableTableWidget()
        self.rules_table.setColumnCount(2)
        self.rules_table.setHorizontalHeaderLabels(["Kode Diagnosa (Regex)", "Kata Kunci Pendukung (Regex)"])
        self.rules_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.layout.addWidget(self.rules_table)
        self.rules_table.rowsMoved.connect(self._update_rules_from_table_order)
        self._populate_table()
        input_grid_layout = QGridLayout()
        input_grid_layout.addWidget(QLabel("Kode Diagnosa (Regex):"), 0, 0)
        self.diag_input = QLineEdit()
        self.diag_input.setPlaceholderText("Misal: i50.*|i11.0|i13.2")
        input_grid_layout.addWidget(self.diag_input, 0, 1)
        input_grid_layout.addWidget(QLabel("Kata Kunci Pendukung:"), 1, 0)
        self.keyword_input = QLineEdit()
        self.keyword_input.setPlaceholderText("Misal: echo|echocardiography")
        input_grid_layout.addWidget(self.keyword_input, 1, 1)
        self.layout.addLayout(input_grid_layout)
        button_row = QHBoxLayout()

        self.add_button = QPushButton("Tambah")
        self.add_button.clicked.connect(self.add_rule)

        self.edit_button = QPushButton("Edit Terpilih")
        self.edit_button.clicked.connect(self.edit_selected_rule)
        self.edit_button.setEnabled(False)

        self.remove_button = QPushButton("Hapus Terpilih")
        self.remove_button.clicked.connect(self.remove_rule)
        self.remove_button.setEnabled(False)

        button_row.addWidget(self.add_button)
        button_row.addWidget(self.edit_button)
        button_row.addWidget(self.remove_button)
        self.layout.addLayout(button_row)
        self.button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        self.layout.addWidget(self.button_box)
        self.setLayout(self.layout)
        self.rules_table.itemSelectionChanged.connect(self.update_buttons_state)
        self.rules_table.itemDoubleClicked.connect(self.edit_selected_rule)
        if parent: self.setStyleSheet(parent.styleSheet())

    def _populate_table(self):
        self.rules_table.setRowCount(0)
        for diag_regex, keyword_regex in self.current_rules:
            row_position = self.rules_table.rowCount()
            self.rules_table.insertRow(row_position)
            self.rules_table.setItem(row_position, 0, QTableWidgetItem(diag_regex))
            self.rules_table.setItem(row_position, 1, QTableWidgetItem(keyword_regex))

    def _update_rules_from_table_order(self, new_order_list):
        self.current_rules = new_order_list

    def update_buttons_state(self):
        has_selection = bool(self.rules_table.selectedItems())
        self.edit_button.setEnabled(has_selection)
        self.remove_button.setEnabled(has_selection)

    def add_rule(self):
        diag_regex = self.diag_input.text().strip()
        keyword_regex = self.keyword_input.text().strip()
        if not diag_regex or not keyword_regex:
            QMessageBox.warning(self, "Input Kosong", "Semua kolom harus diisi."); return
        new_rule_tuple = (diag_regex, keyword_regex)
        if new_rule_tuple in self.current_rules:
            QMessageBox.warning(self, "Duplikat", "Aturan validasi sudah ada dalam daftar."); return
        self.current_rules.append(new_rule_tuple)
        self._populate_table()
        self.diag_input.clear(); self.keyword_input.clear(); self.diag_input.setFocus()
    
    def edit_selected_rule(self):
        selected_items = self.rules_table.selectedItems()
        if selected_items:
            row = selected_items[0].row()
            self.diag_input.setText(self.rules_table.item(row, 0).text())
            self.keyword_input.setText(self.rules_table.item(row, 1).text())
            self.remove_rule(skip_msg=True)
            self.diag_input.setFocus()

    def remove_rule(self, skip_msg=False):
        selected_rows = sorted(list(set(index.row() for index in self.rules_table.selectedIndexes())), reverse=True)
        if selected_rows:
            for row_to_remove in selected_rows:
                if 0 <= row_to_remove < len(self.current_rules): del self.current_rules[row_to_remove]
            self._populate_table()
            if not skip_msg: self.diag_input.clear(); self.keyword_input.clear()
            self.update_buttons_state()
        elif not skip_msg: QMessageBox.warning(self, "Tidak Ada Pilihan", "Pilih aturan yang ingin dihapus.")
    
    def get_updated_rules(self): return self.current_rules


# --- Kelas Utama Aplikasi PDF Verifier ---
class PdfVerifierApp(QWidget):
    def __init__(self):
        super().__init__()
        
        # Panggil metode setup di sini
        self._setup_tesseract_and_dependencies()
        
        self.setWindowTitle("Aplikasi Verifikasi Berkas")
        self.setGeometry(100, 100, 1200, 600)
        self.worker_threads = []
        self.keywords_file = get_resource_path("keywords.json")
        self.rules_file = get_resource_path("rules.json")
        self.list_teks_dicari = []
        self.validation_rules = []
        self.unique_display_headers = []
        
        self.load_keywords()
        self.load_validation_rules()
        self.init_ui()
        self.setStyleSheet(APP_STYLESHEET)
        self.update_table_headers_and_content()
        self.setAcceptDrops(True)
        self.setWindowIcon(QIcon('icon.ico')) 
        self.worker_thread = None
        self.keywords = {}
        self.rules = {}
    
    def _setup_tesseract_and_dependencies(self):
        """Memuat konfigurasi dan memeriksa dependensi yang memerlukan QMessageBox."""
        global pytesseract
        
        # --- KONFIGURASI TESSERACT OCR ---
        config_file = "config.ini"
        config = ConfigParser()

        try:
            config.read(config_file)
            tesseract_path = config.get("Settings", "tesseract_path", fallback="")
        except Exception:
            tesseract_path = ""

        if not os.path.exists(tesseract_path):
            QMessageBox.information(self, "Konfigurasi Tesseract",
                                    "Lokasi Tesseract.exe belum ditemukan. Silakan cari file tesseract.exe.")
            tesseract_path, _ = QFileDialog.getOpenFileName(self, "Cari tesseract.exe", "", "Executable Files (*.exe)")
            
            if tesseract_path:
                if not config.has_section("Settings"):
                    config.add_section("Settings")
                config.set("Settings", "tesseract_path", tesseract_path)
                with open(config_file, 'w') as f:
                    config.write(f)

        try:
            if tesseract_path and os.path.exists(tesseract_path):
                import pytesseract
                pytesseract.pytesseract.tesseract_cmd = tesseract_path
            else:
                raise FileNotFoundError("Tesseract executable not found.")
        except ImportError:
            QMessageBox.warning(self, "Import Error",
                                "Pustaka pytesseract tidak terinstal. Silakan instal dengan 'pip install pytesseract'.")
            pytesseract = None
        except FileNotFoundError as e:
            QMessageBox.warning(self, "Tesseract Not Found",
                                f"Tesseract OCR tidak terinstal atau jalur salah.\n{e}")
            pytesseract = None
        except Exception as e:
            QMessageBox.warning(self, "Konfigurasi Tesseract",
                                f"Gagal mengatur jalur Tesseract OCR.\nError: {e}")
            pytesseract = None
            
        try:
            import fitz # PyMuPDF
        except ImportError:
            QMessageBox.critical(self, "Import Error",
                                 "PyMuPDF library is not installed. Please install it using 'pip install PyMuPDF'.")
            sys.exit(1)

        try:
            from PIL import Image
        except ImportError:
            QMessageBox.critical(self, "Import Error",
                                 "Pillow library is not installed. Please install it using 'pip install Pillow'.")
            sys.exit(1)

    def init_ui(self):
        main_layout = QVBoxLayout()
        header_layout = QHBoxLayout()
        header_layout.addStretch()
        logo_path = get_resource_path("logo.png")
        if os.path.exists(logo_path):
            logo_pixmap = QPixmap(logo_path)
            scaled_logo = logo_pixmap.scaled(100, 100, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
            logo_label = QLabel()
            logo_label.setPixmap(scaled_logo)
            header_layout.addWidget(logo_label)

        self.title_label = QLabel("Verifikasi Berkas PDF")
        self.title_label.setObjectName("title_label")
        header_layout.addWidget(self.title_label)
        header_layout.addStretch()
        main_layout.addLayout(header_layout)
        
        controls_layout = QHBoxLayout()
        controls_layout.addStretch()

        button_layout = QHBoxLayout()
        
        self.select_folder_button = QPushButton("Pilih Folder")
        self.select_folder_button.clicked.connect(self.select_folder)

        self.manage_keywords_button = QPushButton("Atur Kata Kunci")
        self.manage_keywords_button.clicked.connect(self.show_keyword_manager)
        self.manage_rules_button = QPushButton("Atur Aturan Validasi")
        self.manage_rules_button.clicked.connect(self.show_rule_manager)
        self.save_button = QPushButton("Simpan sebagai Excel")
        self.save_button.clicked.connect(self.save_results_to_excel)
        self.save_button.setEnabled(False)

        button_layout.addWidget(self.select_folder_button)
        button_layout.addWidget(self.manage_keywords_button)
        button_layout.addWidget(self.manage_rules_button)
        button_layout.addWidget(self.save_button)
        
        controls_layout.addLayout(button_layout)
        controls_layout.addStretch()

        dpi_layout = QHBoxLayout()
        dpi_label = QLabel("DPI OCR:")
        self.dpi_input = QLineEdit("100")
        self.dpi_input.setFixedWidth(60)
        dpi_layout.addWidget(dpi_label)
        dpi_layout.addWidget(self.dpi_input)
        
        controls_layout.addLayout(dpi_layout)
        controls_layout.addStretch()

        main_layout.addLayout(controls_layout)
        
        self.file_table_widget = DraggableTableWidget()
        self.file_table_widget.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        main_layout.addWidget(self.file_table_widget)
        self.setLayout(main_layout)
    
    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            all_paths = [url.toLocalFile() for url in event.mimeData().urls()]
            valid_paths_found = any(path.lower().endswith('.pdf') or os.path.isdir(path) for path in all_paths)
            if valid_paths_found:
                event.acceptProposedAction()
            else:
                event.ignore()
    
    def dropEvent(self, event):
        if event.mimeData().hasUrls():
            all_paths = [url.toLocalFile() for url in event.mimeData().urls()]
            pdf_paths = []
            for path in all_paths:
                if os.path.isdir(path):
                    for root, _, files in os.walk(path):
                        for file in files:
                            if file.lower().endswith('.pdf'):
                                pdf_paths.append(os.path.join(root, file))
                elif os.path.isfile(path) and path.lower().endswith('.pdf'):
                    pdf_paths.append(path)

            if pdf_paths:
                self.process_selected_pdfs(pdf_paths)
                event.acceptProposedAction()
            else:
                event.ignore()
        else:
            event.ignore()
    
    def load_keywords(self):
        if os.path.exists(self.keywords_file):
            try:
                with open(self.keywords_file, 'r') as f:
                    loaded_data = json.load(f)
                if isinstance(loaded_data, list) and all(isinstance(item, list) and len(item) == 2 and isinstance(item[0], str) and isinstance(item[1], str) for item in loaded_data):
                    self.list_teks_dicari = [tuple(item) for item in loaded_data]
                else:
                    raise ValueError("Format file kata kunci tidak valid.")
            except (json.JSONDecodeError, ValueError) as e:
                QMessageBox.warning(self, "Error Memuat Kata Kunci", f"Gagal memuat kata kunci dari '{self.keywords_file}': {e}\nMenggunakan daftar default kosong."); self.list_teks_dicari = []
        else:
            self.list_teks_dicari = []

    def set_default_rules(self):
        return [
            (r"i50.*|i11.0|i13.2|i13.0", "echo|echocardiography"),
            (r"j44.*", "spirometri"),
            (r"j13|j14.*|j15.*|j16.*|j17.*|j18.*", "thorax"),
            (r"g40.*|g41.*", "eeg"),
            (r"d50.*|d59.*|d62.*|d63.*|d64.*", "transfusi prc")
        ]

    def load_validation_rules(self):
        if os.path.exists(self.rules_file):
            try:
                with open(self.rules_file, 'r') as f:
                    loaded_data = json.load(f)
                if isinstance(loaded_data, list) and all(isinstance(item, list) and len(item) == 2 and isinstance(item[0], str) and isinstance(item[1], str) for item in loaded_data):
                    self.validation_rules = [tuple(item) for item in loaded_data]
                else:
                    raise ValueError("Format file aturan validasi tidak valid. Menggunakan aturan default.")
            except (json.JSONDecodeError, ValueError) as e:
                QMessageBox.warning(self, "Error Memuat Aturan Validasi", f"Gagal memuat aturan validasi dari '{self.rules_file}': {e}\nMenggunakan aturan default."); self.validation_rules = self.set_default_rules()
        else:
            self.validation_rules = self.set_default_rules()
    
    def save_keywords(self):
        try:
            with open(self.keywords_file, 'w') as f:
                json.dump(self.list_teks_dicari, f, indent=4)
        except Exception as e:
            QMessageBox.critical(self, "Error Menyimpan Kata Kunci", f"Gagal menyimpan kata kunci ke '{self.keywords_file}': {e}")

    def save_validation_rules(self):
        try:
            with open(self.rules_file, 'w') as f:
                json.dump(self.validation_rules, f, indent=4)
        except Exception as e:
            QMessageBox.critical(self, "Error Menyimpan Aturan", f"Gagal menyimpan aturan validasi ke '{self.rules_file}': {e}")
    
    def show_keyword_manager(self):
        dialog = KeywordManagerDialog(self.list_teks_dicari, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            updated_keywords = dialog.get_updated_keywords()
            if updated_keywords != self.list_teks_dicari:
                self.list_teks_dicari = updated_keywords
                self.save_keywords()
                self.update_table_headers_and_content()
                QMessageBox.information(self, "Kata Kunci Diperbarui", "Daftar kata kunci telah diperbarui. Silakan proses ulang file PDF.")
            else: QMessageBox.information(self, "Tidak Ada Perubahan", "Tidak ada perubahan pada daftar kata kunci.")
    
    def show_rule_manager(self):
        dialog = ValidationRuleManagerDialog(self.validation_rules, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            updated_rules = dialog.get_updated_rules()
            if updated_rules != self.validation_rules:
                self.validation_rules = updated_rules
                self.save_validation_rules()
                self.update_table_headers_and_content()
                QMessageBox.information(self, "Aturan Diperbarui", "Daftar aturan validasi telah diperbarui. Silakan proses ulang file PDF.")
            else: QMessageBox.information(self, "Tidak Ada Perubahan", "Tidak ada perubahan pada daftar aturan.")

    def update_table_headers_and_content(self):
        ordered_unique_display_texts = []
        seen_display_texts = set()
        
        for _, display_text in self.list_teks_dicari:
            if display_text not in seen_display_texts: 
                ordered_unique_display_texts.append(display_text)
            seen_display_texts.add(display_text)

        self.unique_display_headers = ordered_unique_display_texts
        
        # Perubahan urutan kolom di sini
        num_columns = 1 + 1 + 2 + 1 + 1 + len(self.unique_display_headers)
        self.file_table_widget.setColumnCount(num_columns)
        
        headers = ["NO.", "NAMA\nFILE", "KODE\nDIAGNOSA", "KETERANGAN\nDIAGNOSA", "VALIDASI\nATURAN", "Permintaan\nRanap"] + self.unique_display_headers
        self.file_table_widget.setHorizontalHeaderLabels(headers)
        
        for i in range(self.file_table_widget.columnCount() - 1):
            self.file_table_widget.horizontalHeader().setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)
        
        self.file_table_widget.horizontalHeader().setSectionResizeMode(self.file_table_widget.columnCount() - 1, QHeaderView.ResizeMode.Stretch)
        
        self.file_table_widget.setRowCount(0)
        self.save_button.setEnabled(False)

    def select_pdf_files(self):
        file_dialog = QFileDialog()
        file_paths, _ = file_dialog.getOpenFileNames(self, "Pilih File PDF(s)", "", "PDF Files (*.pdf);;All Files (*)")
        if file_paths:
            self.process_selected_pdfs(file_paths)
        else:
            self.save_button.setEnabled(False)

    def select_folder(self):
        folder_dialog = QFileDialog()
        folder_path = folder_dialog.getExistingDirectory(self, "Pilih Folder")
        if folder_path:
            pdf_paths = []
            for root, _, files in os.walk(folder_path):
                for file in files:
                    if file.lower().endswith('.pdf'):
                        pdf_paths.append(os.path.join(root, file))
            if pdf_paths:
                self.process_selected_pdfs(pdf_paths)
            else:
                self.save_button.setEnabled(False)
        else:
            self.save_button.setEnabled(False)

    def _get_page_text_from_file(self, file_path, page_num, dpi):
        try:
            document = fitz.open(file_path)
            if page_num >= document.page_count:
                document.close()
                return ""
            page = document.load_page(page_num)
            text_from_pdf = page.get_text()
            document.close()
            if text_from_pdf.strip():
                return text_from_pdf
            
            if not pytesseract: return ""
            document = fitz.open(file_path)
            page = document.load_page(page_num)
            zoom = dpi / 72.0
            matrix = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=matrix)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            osd_data = pytesseract.image_to_osd(img)
            rotation_angle_match = re.search(r"Rotate: (\d+)", osd_data)
            if rotation_angle_match:
                angle = int(rotation_angle_match.group(1))
                if angle != 0: img = img.rotate(-angle, expand=True)
            config_ocr = '--psm 3 -l eng+ind'
            text_from_ocr = pytesseract.image_to_string(img, config=config_ocr)
            document.close()
            return text_from_ocr
        except Exception as e:
            return ""

    def process_selected_pdfs(self, file_paths):
        try:
            current_dpi = int(self.dpi_input.text())
            if current_dpi <= 0: raise ValueError("DPI harus bilangan bulat positif.")
        except ValueError as e:
            QMessageBox.warning(self, "Input DPI Tidak Valid", f"DPI harus berupa bilangan bulat positif.\n{e}"); return
        
        self.file_table_widget.setRowCount(0)
        for worker in self.worker_threads: worker.quit(); worker.wait()
        self.worker_threads.clear()
        self.save_button.setEnabled(False)
        self.file_table_widget.setRowCount(len(file_paths))
        
        relevant_files_to_process = []

        for i, file_path in enumerate(file_paths):
            self.file_table_widget.setItem(i, 0, QTableWidgetItem(str(i + 1)))
            self.file_table_widget.setItem(i, 1, QTableWidgetItem(os.path.basename(file_path)))
            
            page_1_text = self._get_page_text_from_file(file_path, 0, current_dpi)
            match_primary = REGEX_EXTRACTION_PRIMARY.search(page_1_text)
            
            code_text = "Tidak Ditemukan"
            description_text = "Tidak Ditemukan"
            is_relevant_code = False
            
            if match_primary:
                extracted_primary_text = match_primary.group(1).strip()
                split_match = REGEX_SPLIT_DIAGNOSA.search(extracted_primary_text)
                if split_match:
                    code_text = split_match.group(2).strip()
                    description_text = split_match.group(3).strip()
                else:
                    description_text = extracted_primary_text
                
                if code_text != "Tidak Ditemukan":
                    code_text_lower = code_text.lower()
                    for diag_keyword_regex, _ in self.validation_rules:
                        if re.search(diag_keyword_regex, code_text_lower):
                            is_relevant_code = True
                            break
            
            # Mendefinisikan ulang indeks kolom berdasarkan urutan baru
            col_kode = 2
            col_keterangan = 3
            col_validation = 4
            col_ringkasan = 5
            col_keywords_start = 6

            kode_item = QTableWidgetItem(code_text)
            kode_item.setForeground(QColor("#C8CDD3") if code_text != "Tidak Ditemukan" else QColor("#636D83"))
            self.file_table_widget.setItem(i, col_kode, kode_item)
            
            keterangan_item = QTableWidgetItem(description_text)
            keterangan_item.setForeground(QColor("#C8CDD3") if description_text != "Tidak Ditemukan" else QColor("#636D83"))
            self.file_table_widget.setItem(i, col_keterangan, keterangan_item)
            
            if not is_relevant_code and self.validation_rules:
                validation_item = QTableWidgetItem("DILEWATI (Kode Diagnosa tidak relevan)")
                validation_item.setForeground(QColor("#636D83"))
                self.file_table_widget.setItem(i, col_validation, validation_item)
                
                status_item_ringkasan = QTableWidgetItem("-")
                status_item_ringkasan.setForeground(QColor("#636D83"))
                self.file_table_widget.setItem(i, col_ringkasan, status_item_ringkasan)

                num_keyword_cols = len(self.unique_display_headers)
                for j in range(col_keywords_start, col_keywords_start + num_keyword_cols):
                    status_item = QTableWidgetItem("-")
                    status_item.setForeground(QColor("#636D83"))
                    self.file_table_widget.setItem(i, j, status_item)
            else:
                num_keyword_cols = len(self.unique_display_headers)
                for j in range(col_validation, col_keywords_start + num_keyword_cols):
                    status_item = QTableWidgetItem("Memproses...")
                    status_item.setForeground(QColor("#00BFFF"))
                    self.file_table_widget.setItem(i, j, status_item)
                
                relevant_files_to_process.append({'path': file_path, 'code': code_text})

        if relevant_files_to_process:
            for file_info in relevant_files_to_process:
                worker = PdfProcessingWorker(file_info['path'], self.list_teks_dicari, current_dpi, self.validation_rules, file_info['code'])
                worker.finished.connect(self.on_processing_finished)
                worker.keyword_found_signal.connect(self.on_keyword_found)
                self.worker_threads.append(worker)
                worker.start()
        else:
            self.save_button.setEnabled(True)
        
        self.file_table_widget.resizeColumnsToContents()
        self.file_table_widget.resizeRowsToContents()

    def on_keyword_found(self, pdf_path, display_text, page_number):
        row = -1
        for i in range(self.file_table_widget.rowCount()):
            if self.file_table_widget.item(i, 1).text() == os.path.basename(pdf_path):
                row = i
                break
        if row == -1: return

        # Indeks baru
        col_keywords_start = 6
        col = -1
        for i in range(col_keywords_start, col_keywords_start + len(self.unique_display_headers)):
            header = self.file_table_widget.horizontalHeaderItem(i).text().replace('\n', ' ')
            if header == display_text:
                col = i
                break
        if col == -1: return

        item_text = f"✓ (hal. {page_number})"
        item = QTableWidgetItem(item_text)
        item.setForeground(QColor("#00FF00"))
        self.file_table_widget.setItem(row, col, item)

    def on_processing_finished(self, pdf_path, results, error_message, found_page_numbers, found_ringkasan_keyword):
        row = -1
        for i in range(self.file_table_widget.rowCount()):
            if self.file_table_widget.item(i, 1).text() == os.path.basename(pdf_path): row = i; break
        if row == -1: return

        # Indeks baru
        col_kode = 2
        col_validation = 4
        col_ringkasan = 5
        col_keywords_start = 6
        
        ringkasan_item = QTableWidgetItem("✓" if found_ringkasan_keyword else "✗")
        ringkasan_item.setForeground(QColor("#00FF00") if found_ringkasan_keyword else QColor("#FF0000"))
        self.file_table_widget.setItem(row, col_ringkasan, ringkasan_item)
        
        num_keyword_cols = len(self.unique_display_headers)
        for col_idx, display_text in enumerate(self.unique_display_headers, col_keywords_start):
            current_item = self.file_table_widget.item(row, col_idx)
            if not current_item or current_item.text() == "Memproses...":
                item_text = "✗"
                item = QTableWidgetItem(item_text)
                item.setForeground(QColor("#FF0000"))
                self.file_table_widget.setItem(row, col_idx, item)

        validation_status_list = []
        code_text = self.file_table_widget.item(row, col_kode).text()
        
        if code_text != "Tidak Ditemukan":
            code_text_lower = code_text.lower()
            for diag_keyword_regex, must_have_keyword in self.validation_rules:
                if re.search(diag_keyword_regex, code_text_lower):
                    page_found = found_page_numbers.get(must_have_keyword.upper(), -1)
                    if page_found != -1:
                        validation_status_list.append(f"AMAN (di hlm. {page_found})")
                        
                        col = -1
                        for i in range(col_keywords_start, col_keywords_start + num_keyword_cols):
                            header = self.file_table_widget.horizontalHeaderItem(i).text().replace('\n', ' ').strip()
                            if header.lower() == must_have_keyword.lower():
                                col = i
                                break
                        if col != -1:
                            item_text = f"✓ (hal. {page_found})"
                            item = QTableWidgetItem(item_text)
                            item.setForeground(QColor("#00FF00"))
                            self.file_table_widget.setItem(row, col, item)
                    else:
                        validation_status_list.append(f"TIDAK AMAN (harus ada '{must_have_keyword}')")
        
        if not validation_status_list:
            validation_item = QTableWidgetItem("LULUS")
            validation_color = QColor("#00FF00")
        else:
            validation_status = ", ".join(validation_status_list)
            validation_item = QTableWidgetItem(validation_status)
            validation_color = QColor("#FF0000") if "TIDAK LULUS" in validation_status else QColor("#00FF00")

        validation_item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | WORD_WRAP_FLAG)
        validation_item.setForeground(validation_color)
        self.file_table_widget.setItem(row, col_validation, validation_item)
        
        self.file_table_widget.resizeColumnsToContents()
        self.file_table_widget.resizeRowsToContents()

        all_workers_finished = all(not worker.isRunning() for worker in self.worker_threads)
        if all_workers_finished:
            self.save_button.setEnabled(True)

    def save_results_to_excel(self):
        file_dialog = QFileDialog()
        excel_path, _ = file_dialog.getSaveFileName(self, "Simpan Hasil", "Hasil_Verifikasi.xlsx", "Excel Files (*.xlsx)")
        if not excel_path:
            return

        try:
            from openpyxl import load_workbook
            import os
            
            # --- Periksa keberadaan file ---
            file_exists = os.path.exists(excel_path)
            
            if file_exists:
                # Jika file ada, muat dan dapatkan worksheet aktif
                wb = load_workbook(excel_path)
                ws = wb.active
                start_row = ws.max_row + 1
            else:
                # Jika file tidak ada, buat workbook baru dan tambahkan header
                wb = Workbook()
                ws = wb.active
                ws.title = "Hasil Verifikasi"
                start_row = 1
                
                # --- MENAMBAH HEADER DAN FORMATNYA (HANYA UNTUK FILE BARU) ---
                headers = [self.file_table_widget.horizontalHeaderItem(i).text().replace('\n', ' ') for i in range(self.file_table_widget.columnCount())]
                ws.append(headers)
                
                header_font = ExcelFont(bold=True, color="000000")
                header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                header_border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
                
                for col_idx, header_text in enumerate(headers, 1):
                    cell = ws.cell(row=start_row, column=col_idx)
                    cell.value = header_text
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = header_border
                start_row += 1

            # --- MENAMBAHKAN DATA BARU ---
            for row in range(self.file_table_widget.rowCount()):
                row_data = [self.file_table_widget.item(row, col).text() if self.file_table_widget.item(row, col) else "" for col in range(self.file_table_widget.columnCount())]
                ws.append(row_data)

            # --- MENYESUAIKAN LEBAR KOLOM ---
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                             if '\n' in str(cell.value):
                                 cell.alignment = Alignment(wrap_text=True)
                             content_length = len(max(str(cell.value).split('\n'), key=len))
                             if content_length > max_length: max_length = content_length
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            # --- Simpan buku kerja ---
            wb.save(excel_path)
            QMessageBox.information(self, "Simpan Berhasil", f"Hasil berhasil disimpan ke:\n{excel_path}")

        except Exception as e:
            QMessageBox.critical(self, "Error Menyimpan", f"Gagal menyimpan file Excel: {e}")
            header_font = ExcelFont(bold=True, color="000000")
            header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            header_border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
            
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border
            
            for row in range(self.file_table_widget.rowCount()):
                row_data = [self.file_table_widget.item(row, col).text() if self.file_table_widget.item(row, col) else "" for col in range(self.file_table_widget.columnCount())]
                ws.append(row_data)
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                             if '\n' in str(cell.value):
                                 cell.alignment = Alignment(wrap_text=True)
                             content_length = len(max(str(cell.value).split('\n'), key=len))
                             if content_length > max_length: max_length = content_length
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(excel_path)
            QMessageBox.information(self, "Simpan Berhasil", f"Hasil berhasil disimpan ke:\n{excel_path}")
        except Exception as e:
            QMessageBox.critical(self, "Error Menyimpan", f"Gagal menyimpan file Excel: {e}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = PdfVerifierApp()
    window.show()
    sys.exit(app.exec())
