import sys
import os
import re
import json

from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QFileDialog, QLabel, QTableWidget, QTableWidgetItem, QMessageBox,
    QHeaderView, QDialog, QListWidget, QLineEdit, QDialogButtonBox,
    QGridLayout, QCheckBox
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QMimeData, QDataStream, QIODevice
from PyQt6.QtGui import QColor, QFont

import fitz # PyMuPDF
import pytesseract
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font as ExcelFont, PatternFill, Alignment, Side, Border


# --- KONFIGURASI TESSERACT OCR ---
try:
    # >>>>>> PENTING: SESUAIKAN PATH INI DENGAN LOKASI TESSERACT.EXE ANDA <<<<<<
    # Contoh untuk Windows:
    # pytesseract.pytesseract.tesseract_cmd = r'C:\Users\NAMA_ANDA\AppData\Local\Programs\Tesseract-OCR\tesseract.exe'
    pytesseract.pytesseract.tesseract_cmd = r'C:\Users\IJP-INDI\AppData\Local\Programs\Tesseract-OCR\tesseract.exe'
except AttributeError:
    pass
except Exception as e:
    print(f"Peringatan: Gagal mengatur path Tesseract OCR. Pastikan Tesseract terinstal dan path benar. Error: {e}")
    QMessageBox.warning(None, "Konfigurasi Tesseract",
                        f"Gagal mengatur path Tesseract OCR. Pastikan Tesseract terinstal dan path-nya benar.\nError: {e}")


# --- KONSTANTA PENTING ---
# Kata kunci untuk menghentikan pemindaian dari bawah ke atas
STOP_KEYWORD = "SEP"
# Kata kunci yang teksnya akan diekstrak untuk kolom terakhir (misal: "Diagnosa Akhir")
EXTRACTION_KEYWORD = "Diagnosa Akhir"

# Regex untuk mengekstrak teks setelah EXTRACTION_KEYWORD
REGEX_EXTRACTION = re.compile(rf"{EXTRACTION_KEYWORD}\s*[:;,-]?\s*(.*?)(?:\n|$)", re.IGNORECASE | re.DOTALL)


# --- STYLESHEET TEMA ABU-ABU GELAP DENGAN AKSEN TEAL ---
APP_STYLESHEET = """
QWidget {
    background-color: #333333; /* Abu-abu gelap sangat tua */
    color: #C8CDD3; /* Abu-abu terang untuk teks umum */
    font-family: 'Segoe UI', 'Roboto', sans-serif;
    font-size: 12px;
}

QLabel#title_label {
    color: #00BFFF; /* Deep Sky Blue sebagai aksen terang untuk judul */
    font-size: 24px;
    font-weight: bold;
    margin-bottom: 20px;
    qproperty-alignment: 'AlignHCenter';
}

QPushButton {
    background-color: #30363D; /* Abu-abu sangat gelap untuk background tombol */
    color: #FFFFFF; /* Teks putih untuk tombol */
    border: 1px solid #00BFFF; /* Border Deep Sky Blue */
    border-radius: 8px; /* Sudut membulat */
    padding: 8px 15px;
    font-weight: bold;
    text-transform: uppercase;
}

QPushButton:hover {
    background-color: #3B424A; /* Sedikit lebih terang saat hover */
    border: 1px solid #4AC8FF; /* Border Deep Sky Blue yang sedikit lebih terang saat hover */
}

QPushButton:pressed {
    background-color: #00BFFF; /* Warna aksen saat ditekan */
    color: #21252B; /* Teks gelap saat ditekan */
}

QPushButton:disabled {
    background-color: #282C34;
    color: #636D83;
    border: 1px solid #282C34;
}

QTableWidget {
    background-color: #21252B; /* Background tabel sama dengan background utama */
    alternate-background-color: #2A2F36; /* Baris alternatif sedikit lebih gelap */
    color: #C8CDD3; /* Warna teks tabel */
    gridline-color: #30363D; /* Garis grid abu-abu gelap yang halus */
    border: 1px solid #00BFFF; /* Border tabel Deep Sky Blue */
    selection-background-color: #00BFFF; /* Warna seleksi Deep Sky Blue */
    selection-color: #21252B; /* Teks seleksi gelap */
    border-radius: 8px;
}

QTableWidget::item {
    padding: 4px;
}

QHeaderView::section {
    background-color: #30363D; /* Header abu-abu gelap */
    color: #00BFFF; /* Teks header Deep Sky Blue */
    padding: 6px;
    border: 1px solid #21252B; /* Border menyatu dengan background tabel */
    font-weight: bold;
    text-transform: uppercase;
}

QHeaderView::section:horizontal {
    border-top: 0px solid #21252B;
    border-bottom: 1px solid #00BFFF; /* Garis bawah header Deep Sky Blue */
}

QMessageBox {
    background-color: #21252B;
    color: #C8CDD3;
    font-size: 12px;
}

QMessageBox QPushButton {
    background-color: #00BFFF;
    color: #21252B;
    border: none;
    padding: 7px 12px;
    border-radius: 5px;
    font-weight: bold;
}

QMessageBox QPushButton:hover {
    background-color: #4AC8FF;
    color: #21252B;
}

QLineEdit {
    background-color: #30363D;
    color: #C8CDD3;
    border: 1px solid #00BFFF;
    border-radius: 5px;
    padding: 5px;
}

QListWidget {
    background-color: #2A2F36;
    color: #C8CDD3;
    border: 1px solid #30363D;
    border-radius: 5px;
    selection-background-color: #00BFFF;
    selection-color: #21252B;
}
QListWidget::item {
    padding: 4px;
}
QLabel { /* Pastikan semua QLabel mendapatkan warna teks yang benar */
    color: #C8CDD3;
}
QCheckBox {
    color: #C8CDD3;
    spacing: 5px;
}
QCheckBox::indicator {
    width: 16px;
    height: 16px;
    border-radius: 3px;
    border: 1px solid #00BFFF;
    background-color: #30363D;
}
QCheckBox::indicator:checked {
    background-color: #00BFFF;
    border: 1px solid #4AC8FF;
}
QCheckBox::indicator:hover {
    border: 1px solid #4AC8FF;
}
"""

# --- Kelas Worker untuk Memproses PDF (Berjalan di Thread Terpisah) ---
class PdfProcessingWorker(QThread):
    finished = pyqtSignal(str, dict, str, str)

    def __init__(self, pdf_path, texts_to_find_tuples, dpi, **kwargs):
        super().__init__()
        self.pdf_path = pdf_path
        self.texts_to_find_tuples = texts_to_find_tuples # List of (search_text, display_text)
        self.dpi = dpi
        self.process_from_bottom = kwargs.get('process_from_bottom', False)
        self.stop_on_keyword_flag = kwargs.get('stop_on_keyword_flag', False)

    def run(self):
        # Group search texts by their display text
        grouped_search_texts = {}
        for search_text, display_text in self.texts_to_find_tuples:
            grouped_search_texts.setdefault(display_text, []).append(search_text)

        # Initialize results for each unique display text
        results = {display_text: False for display_text in grouped_search_texts.keys()}
        
        error_message = ""
        extracted_final_text = ""
        document = None

        found_stop_keyword_in_page = False

        try:
            document = fitz.open(self.pdf_path)
            full_text_content = ""

            if self.process_from_bottom:
                page_numbers_to_process = range(document.page_count - 1, -1, -1)
                print(f"DEBUG: Memproses '{os.path.basename(self.pdf_path)}' dari bawah. Halaman yang akan dicoba diproses: {list(page_numbers_to_process)}")
            else:
                page_numbers_to_process = range(document.page_count)
                print(f"DEBUG: Memproses '{os.path.basename(self.pdf_path)}' dari atas. Halaman yang akan dicoba diproses: {list(page_numbers_to_process)}")


            for page_num in page_numbers_to_process:
                if self.stop_on_keyword_flag and found_stop_keyword_in_page:
                    print(f"DEBUG: '{STOP_KEYWORD}' sudah ditemukan di halaman sebelumnya, berhenti memproses halaman {page_num + 1} dan seterusnya.")
                    break

                page = document.load_page(page_num)
                page_text_from_pdf = ""
                page_text_from_ocr = ""
                pix = None
                img = None

                try:
                    page_text_from_pdf = page.get_text()

                    if not page_text_from_pdf.strip():
                        print(f"DEBUG: Halaman {page_num + 1} di '{os.path.basename(self.pdf_path)}' tidak memiliki teks langsung, mencoba OCR dengan DPI {self.dpi}.")

                        zoom = self.dpi / 72.0
                        matrix = fitz.Matrix(zoom, zoom)
                        pix = page.get_pixmap(matrix=matrix)

                        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

                        try:
                            osd_data = pytesseract.image_to_osd(img)
                            rotation_angle_match = re.search(r"Rotate: (\d+)", osd_data)

                            if rotation_angle_match:
                                angle = int(rotation_angle_match.group(1))
                                if angle != 0:
                                    img = img.rotate(-angle, expand=True)
                                    print(f"DEBUG: Halaman {page_num + 1} diputar {angle} derajat.")

                            config_ocr = '--psm 3 -l eng+ind'
                            page_text_from_ocr = pytesseract.image_to_string(img, config=config_ocr)

                        except pytesseract.TesseractNotFoundError:
                            error_message = "Mesin Tesseract OCR tidak ditemukan. Pastikan sudah terinstal dan path-nya benar, serta variabel TESSDATA_PREFIX telah diatur jika perlu."
                            print(f"ERROR: {error_message}")
                            break
                        except Exception as e:
                            if "Too few characters" in str(e) or "Error during processing." in str(e) or "Invalid resolution" in str(e):
                                print(f"WARNING: Error OCR yang diabaikan di halaman {page_num + 1} (kemungkinan halaman kosong/noise): {e}")
                                page_text_from_ocr = ""
                            else:
                                error_message = f"Error serius saat OCR di halaman {page_num + 1} (gambar): {e}"
                                print(f"ERROR: {error_message}")
                                break

                        finally:
                            if pix:
                                del pix
                                pix = None

                    current_page_text = page_text_from_pdf if page_text_from_pdf.strip() else page_text_from_ocr

                    if re.sub(r'\s+', '', current_page_text) == '':
                        print(f"DEBUG: Halaman {page_num + 1} di '{os.path.basename(self.pdf_path)}' dianggap kosong dan diabaikan.")
                        continue

                    else:
                        if self.process_from_bottom:
                            full_text_content = current_page_text + "\n" + full_text_content
                        else:
                            full_text_content += current_page_text + "\n"

                        if STOP_KEYWORD.lower() in current_page_text.lower():
                            found_stop_keyword_in_page = True
                            print(f"DEBUG: '{STOP_KEYWORD}' DITEMUKAN DI HALAMAN {page_num + 1}! Menyiapkan untuk berhenti.")


                except Exception as e:
                    error_message = f"Terjadi kesalahan saat memproses halaman {page_num + 1}: {e}"
                    print(f"ERROR: {error_message}")
                    break

            if not error_message:
                # Process based on grouped search texts
                for display_text, search_texts_list in grouped_search_texts.items():
                    found_any_for_display = False
                    for search_text in search_texts_list:
                        if search_text.lower() in full_text_content.lower():
                            found_any_for_display = True
                            print(f"DEBUG: Keyword pencarian '{search_text}' (untuk tampilan '{display_text}') DITEMUKAN.")
                            break # Found at least one for this display_text
                    results[display_text] = found_any_for_display
                    if not found_any_for_display:
                         print(f"DEBUG: TIDAK ADA keyword pencarian ditemukan untuk tampilan '{display_text}'.")

                match = REGEX_EXTRACTION.search(full_text_content)
                if match:
                    extracted_final_text = match.group(1).strip()
                    print(f"DEBUG: '{EXTRACTION_KEYWORD}' ditemukan. Teks diekstrak: '{extracted_final_text}'")
                else:
                    extracted_final_text = "Tidak Ditemukan"
                    print(f"DEBUG: '{EXTRACTION_KEYWORD}' TIDAK DITEMUKAN.")
            else:
                extracted_final_text = f"Error: {error_message}"

        except fitz.FileNotFoundError:
            error_message = "File PDF tidak ditemukan."
            extracted_final_text = f"Error: {error_message}"
        except Exception as e:
            error_message = f"Terjadi kesalahan umum saat membuka/memproses PDF: {e}"
            extracted_final_text = f"Error: {e}"

        finally:
            if document:
                document.close()

        self.finished.emit(self.pdf_path, results, error_message, extracted_final_text)


# --- Kelas QTableWidget Kustom untuk Drag & Drop ---
class DraggableTableWidget(QTableWidget):
    rowsMoved = pyqtSignal(list) # Sinyal untuk memberitahu parent saat baris dipindahkan

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setDragDropMode(QTableWidget.DragDropMode.InternalMove)
        self.setDragEnabled(True)
        self.setAcceptDrops(True)
        self.setDropIndicatorShown(True)
        self.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection) # Memungkinkan seleksi multiple

    def dropEvent(self, event):
        if event.source() == self and (event.dropAction() == Qt.DropAction.MoveAction or event.proposedAction() == Qt.DropAction.MoveAction):
            selected_rows_indices = sorted([index.row() for index in self.selectedIndexes()])
            if not selected_rows_indices:
                event.ignore()
                return

            moved_items_data = []
            for r_idx in selected_rows_indices:
                search_text = self.item(r_idx, 0).text()
                display_text = self.item(r_idx, 1).text()
                moved_items_data.append((search_text, display_text))

            drop_row = self.indexAt(event.position().toPoint()).row()
            if drop_row == -1:
                drop_row = self.rowCount()

            self.blockSignals(True)

            for r_idx in sorted(selected_rows_indices, reverse=True):
                self.removeRow(r_idx)

            for i, item_data in enumerate(moved_items_data):
                insert_at = drop_row + i
                self.insertRow(insert_at)
                self.setItem(insert_at, 0, QTableWidgetItem(item_data[0]))
                self.setItem(insert_at, 1, QTableWidgetItem(item_data[1]))

            self.blockSignals(False)

            new_keywords_order = []
            for r in range(self.rowCount()):
                search_text = self.item(r, 0).text()
                display_text = self.item(r, 1).text()
                new_keywords_order.append((search_text, display_text))
            self.rowsMoved.emit(new_keywords_order)

            event.acceptProposedAction()
        else:
            super().dropEvent(event)


# --- Kelas Dialog untuk Mengatur Kata Kunci ---
class KeywordManagerDialog(QDialog):
    def __init__(self, current_keywords_tuples, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Atur Kata Kunci Pencarian")
        self.setGeometry(200, 200, 600, 450)

        self.current_keywords = list(current_keywords_tuples)

        self.layout = QVBoxLayout()

        self.keyword_table = DraggableTableWidget()
        self.keyword_table.setColumnCount(2)
        self.keyword_table.setHorizontalHeaderLabels(["Teks Pencarian (Internal)", "Teks Tampilan (GUI)"])
        self.keyword_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.layout.addWidget(self.keyword_table)

        self.keyword_table.rowsMoved.connect(self._update_keywords_from_table_order)

        self._populate_table()

        input_grid_layout = QGridLayout()
        input_grid_layout.addWidget(QLabel("Teks Pencarian:"), 0, 0)
        self.search_input_field = QLineEdit()
        self.search_input_field.setPlaceholderText("Teks yang akan dicari di PDF (misal: 'Prosedur Non-Bedah')")
        input_grid_layout.addWidget(self.search_input_field, 0, 1)

        input_grid_layout.addWidget(QLabel("Teks Tampilan:"), 1, 0)
        self.display_input_field = QLineEdit()
        self.display_input_field.setPlaceholderText("Teks yang akan ditampilkan di tabel utama (misal: 'Non-Bedah')")
        input_grid_layout.addWidget(self.display_input_field, 1, 1)
        self.layout.addLayout(input_grid_layout)

        button_row = QHBoxLayout()
        self.add_button = QPushButton("Tambah")
        self.add_button.clicked.connect(self.add_keyword)

        self.edit_button = QPushButton("Edit Terpilih")
        self.edit_button.clicked.connect(self.edit_selected_keyword)
        self.edit_button.setEnabled(False)

        self.remove_button = QPushButton("Hapus Terpilih")
        self.remove_button.clicked.connect(self.remove_keyword)
        self.remove_button.setEnabled(False)

        button_row.addWidget(self.add_button)
        button_row.addWidget(self.edit_button)
        button_row.addWidget(self.remove_button)
        self.layout.addLayout(button_row)

        self.button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        self.layout.addWidget(self.button_box)

        self.setLayout(self.layout)

        self.keyword_table.itemSelectionChanged.connect(self.update_buttons_state)
        self.keyword_table.itemDoubleClicked.connect(self.edit_selected_keyword)

        if parent:
            self.setStyleSheet(parent.styleSheet())
            self.keyword_table.setStyleSheet("""
                QTableWidget {
                    background-color: #2A2F36;
                    color: #C8CDD3;
                    border: 1px solid #30363D;
                    border-radius: 5px;
                    selection-background-color: #00BFFF;
                    selection-color: #21252B;
                }
                QTableWidget::item {
                    padding: 4px;
                }
                QHeaderView::section {
                    background-color: #30363D;
                    color: #00BFFF;
                    padding: 6px;
                    border: 1px solid #21252B;
                    font-weight: bold;
                    text-transform: uppercase;
                }
                QHeaderView::section:horizontal {
                    border-top: 0px solid #21252B;
                    border-bottom: 1px solid #00BFFF;
                }
            """)
            self.search_input_field.setStyleSheet("""
                QLineEdit {
                    background-color: #30363D;
                    color: #C8CDD3;
                    border: 1px solid #00BFFF;
                    border-radius: 5px;
                    padding: 5px;
                }
            """)
            self.display_input_field.setStyleSheet("""
                QLineEdit {
                    background-color: #30363D;
                    color: #C8CDD3;
                    border: 1px solid #00BFFF;
                    border-radius: 5px;
                    padding: 5px;
                }
            """)
            for label in self.findChildren(QLabel):
                label.setStyleSheet("QLabel { color: #C8CDD3; }")

    def _populate_table(self):
        self.keyword_table.setRowCount(0)
        for search_text, display_text in self.current_keywords:
            row_position = self.keyword_table.rowCount()
            self.keyword_table.insertRow(row_position)
            self.keyword_table.setItem(row_position, 0, QTableWidgetItem(search_text))
            self.keyword_table.setItem(row_position, 1, QTableWidgetItem(display_text))

    def _update_keywords_from_table_order(self, new_order_list):
        self.current_keywords = new_order_list

    def update_buttons_state(self):
        has_selection = bool(self.keyword_table.selectedItems())
        self.edit_button.setEnabled(has_selection)
        self.remove_button.setEnabled(has_selection)

    def add_keyword(self):
        search_text = self.search_input_field.text().strip()
        display_text = self.display_input_field.text().strip()

        if not search_text or not display_text:
            QMessageBox.warning(self, "Input Kosong", "Teks Pencarian dan Teks Tampilan tidak boleh kosong.")
            return

        new_keyword_tuple = (search_text, display_text)

        # Check for duplicate search_text (internal text) as each row is unique by search_text
        if any(kw[0].lower() == search_text.lower() for kw in self.current_keywords):
            QMessageBox.warning(self, "Duplikat", "Teks Pencarian sudah ada dalam daftar.")
            return

        self.current_keywords.append(new_keyword_tuple)
        self._populate_table()
        self.search_input_field.clear()
        self.display_input_field.clear()
        self.search_input_field.setFocus()

    def edit_selected_keyword(self):
        selected_items = self.keyword_table.selectedItems()
        if selected_items:
            row = selected_items[0].row()

            self.search_input_field.setText(self.keyword_table.item(row, 0).text())
            self.display_input_field.setText(self.keyword_table.item(row, 1).text())

            # Remove the old entry
            self.remove_keyword(skip_msg=True)
            self.search_input_field.setFocus()

    def remove_keyword(self, skip_msg=False):
        selected_rows = sorted(list(set(index.row() for index in self.keyword_table.selectedIndexes())), reverse=True)
        if selected_rows:
            for row_to_remove in selected_rows:
                if 0 <= row_to_remove < len(self.current_keywords):
                    del self.current_keywords[row_to_remove]
            
            self._populate_table()

            if not skip_msg:
                self.search_input_field.clear()
                self.display_input_field.clear()
            self.update_buttons_state()
        elif not skip_msg:
            QMessageBox.warning(self, "Tidak Ada Pilihan", "Pilih kata kunci yang ingin dihapus.")

    def get_updated_keywords(self):
        return self.current_keywords

# --- Kelas Utama Aplikasi PDF Verifier ---
class PdfVerifierApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Aplikasi Verifikasi Berkas")
        self.setGeometry(100, 100, 1200, 600)

        self.worker_threads = []
        self.keywords_file = "keywords.json"
        self.list_teks_dicari = [] # Stores (search_text, display_text)
        self.unique_display_headers = [] # Stores unique display texts for table headers

        self.load_keywords()
        self.init_ui() # Baris ini memanggil init_ui
        self.setStyleSheet(APP_STYLESHEET)
        self.update_table_headers_and_content() # Call initially to set up headers

        # Enable drag and drop for the main window
        self.setAcceptDrops(True)

    def init_ui(self): # Definisikan init_ui di sini, sejajar dengan __init__
        main_layout = QVBoxLayout()
        top_layout = QHBoxLayout()

        self.title_label = QLabel("Verifikasi Berkas PDF")
        self.title_label.setObjectName("title_label")
        main_layout.addWidget(self.title_label, alignment=Qt.AlignmentFlag.AlignCenter)

        self.select_button = QPushButton("Pilih File PDF(s)")
        self.select_button.clicked.connect(self.select_pdf_files)
        top_layout.addWidget(self.select_button)

        self.manage_keywords_button = QPushButton("Atur Kata Kunci")
        self.manage_keywords_button.clicked.connect(self.show_keyword_manager)
        top_layout.addWidget(self.manage_keywords_button)

        dpi_layout = QHBoxLayout()
        dpi_layout.addWidget(QLabel("DPI OCR (misal: 150, 300, 600):"))
        self.dpi_input = QLineEdit("300")
        self.dpi_input.setFixedWidth(60)
        self.dpi_input.setAlignment(Qt.AlignmentFlag.AlignCenter)
        dpi_layout.addWidget(self.dpi_input)
        dpi_layout.addStretch()
        top_layout.addLayout(dpi_layout)

        self.scan_from_bottom_checkbox = QCheckBox(f"Pindai dari Halaman Terakhir (hingga '{STOP_KEYWORD}' ditemukan)")
        self.scan_from_bottom_checkbox.setChecked(True)
        top_layout.addWidget(self.scan_from_bottom_checkbox)

        self.save_button = QPushButton("Simpan sebagai Excel")
        self.save_button.clicked.connect(self.save_results_to_excel)
        self.save_button.setEnabled(False)
        top_layout.addWidget(self.save_button)

        self.status_label = QLabel("Siap untuk verifikasi.")
        top_layout.addWidget(self.status_label)

        main_layout.addLayout(top_layout)

        self.file_table_widget = QTableWidget()
        self.file_table_widget.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.file_table_widget.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        main_layout.addWidget(self.file_table_widget)

        self.setLayout(main_layout)

    def dragEnterEvent(self, event):
        """Handle drag enter event to check for PDF files."""
        if event.mimeData().hasUrls():
            pdf_found = False
            for url in event.mimeData().urls():
                if url.isLocalFile() and url.toLocalFile().lower().endswith('.pdf'):
                    pdf_found = True
                    break
            if pdf_found:
                event.acceptProposedAction()
            else:
                event.ignore()
        else:
            event.ignore()

    def dragMoveEvent(self, event):
        """Handle drag move event for visual feedback."""
        if event.mimeData().hasUrls():
            pdf_found = False
            for url in event.mimeData().urls():
                if url.isLocalFile() and url.toLocalFile().lower().endswith('.pdf'):
                    pdf_found = True
                    break
            if pdf_found:
                event.acceptProposedAction()
            else:
                event.ignore()
        else:
            event.ignore()

    def dropEvent(self, event):
        """Handle drop event to process PDF files."""
        if event.mimeData().hasUrls():
            pdf_paths = []
            for url in event.mimeData().urls():
                if url.isLocalFile() and url.toLocalFile().lower().endswith('.pdf'):
                    pdf_paths.append(url.toLocalFile())
            
            if pdf_paths:
                self.process_selected_pdfs(pdf_paths)
                event.acceptProposedAction()
            else:
                event.ignore()
        else:
            event.ignore()

    def load_keywords(self):
        if os.path.exists(self.keywords_file):
            try:
                with open(self.keywords_file, 'r') as f:
                    loaded_data = json.load(f)

                if isinstance(loaded_data, list) and \
                    all(isinstance(item, list) and len(item) == 2 and
                        isinstance(item[0], str) and isinstance(item[1], str)
                        for item in loaded_data):
                    self.list_teks_dicari = [tuple(item) for item in loaded_data]
                else:
                    raise ValueError("Format file kata kunci tidak valid. Harus berupa list of [teks_pencarian, teks_tampilan].")
            except (json.JSONDecodeError, ValueError) as e:
                QMessageBox.warning(self, "Error Memuat Kata Kunci", f"Gagal memuat kata kunci dari '{self.keywords_file}': {e}\nMenggunakan daftar default kosong.")
                self.set_default_keywords()
        else:
            self.set_default_keywords()

    def set_default_keywords(self):
        self.list_teks_dicari = []

    def save_keywords(self):
        try:
            with open(self.keywords_file, 'w') as f:
                json.dump(self.list_teks_dicari, f, indent=4)
        except Exception as e:
            QMessageBox.critical(self, "Error Menyimpan Kata Kunci", f"Gagal menyimpan kata kunci ke '{self.keywords_file}': {e}")

    def show_keyword_manager(self):
        dialog = KeywordManagerDialog(self.list_teks_dicari, self)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            updated_keywords = dialog.get_updated_keywords()
            if updated_keywords != self.list_teks_dicari:
                self.list_teks_dicari = updated_keywords
                self.save_keywords()
                self.update_table_headers_and_content() # Re-generate columns based on updated unique display texts
                QMessageBox.information(self, "Kata Kunci Diperbarui", "Daftar kata kunci telah diperbarui. Silakan proses ulang file PDF untuk menerapkan perubahan.")
            else:
                QMessageBox.information(self, "Tidak Ada Perubahan", "Tidak ada perubahan pada daftar kata kunci.")
        else:
            QMessageBox.information(self, "Dibatalkan", "Pengaturan kata kunci dibatalkan.")

    def update_table_headers_and_content(self):
        # *** PERBAIKAN DI SINI: Ekstrak unique_display_headers sambil mempertahankan urutan ***
        ordered_unique_display_texts = []
        seen_display_texts = set()
        for search_text, display_text in self.list_teks_dicari:
            if display_text not in seen_display_texts:
                ordered_unique_display_texts.append(display_text)
                seen_display_texts.add(display_text)
        self.unique_display_headers = ordered_unique_display_texts
        # ***********************************************************************************

        num_columns = 1 + 1 + len(self.unique_display_headers) + 1 # No. + File Name + Unique Keywords + Extra Extraction
        self.file_table_widget.setColumnCount(num_columns)

        headers = ["No."] + ["Nama File"] + self.unique_display_headers + [EXTRACTION_KEYWORD]
        self.file_table_widget.setHorizontalHeaderLabels(headers)

        self.file_table_widget.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.file_table_widget.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        for i in range(2, num_columns - 1): # Keyword columns
            self.file_table_widget.horizontalHeader().setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)
        self.file_table_widget.horizontalHeader().setSectionResizeMode(num_columns - 1, QHeaderView.ResizeMode.Stretch) # Last extraction column

        self.file_table_widget.setRowCount(0)
        self.save_button.setEnabled(False)
        self.status_label.setText("Header tabel diperbarui. Siap untuk verifikasi file baru.")

    def select_pdf_files(self):
        """Opens file dialog and passes selected files for processing."""
        file_dialog = QFileDialog()
        file_paths, _ = file_dialog.getOpenFileNames(
            self, "Pilih File PDF(s)", "", "PDF Files (*.pdf);;All Files (*)"
        )
        if file_paths:
            self.process_selected_pdfs(file_paths)
        else:
            self.status_label.setText("Pemilihan file dibatalkan.")
            self.save_button.setEnabled(False)

    def process_selected_pdfs(self, file_paths):
        """Centralized method to process a list of PDF file paths."""
        # Validate DPI input
        try:
            current_dpi = int(self.dpi_input.text())
            if current_dpi <= 0:
                raise ValueError("DPI harus bilangan bulat positif.")
        except ValueError as e:
            QMessageBox.warning(self, "Input DPI Tidak Valid", f"DPI harus berupa bilangan bulat positif.\n{e}")
            return

        # Clear previous results and stop active workers
        self.file_table_widget.setRowCount(0)
        for worker in self.worker_threads:
            worker.quit()
            worker.wait()
        self.worker_threads.clear()

        self.status_label.setText(f"Memproses {len(file_paths)} file, mohon tunggu...")
        self.save_button.setEnabled(False)

        process_from_bottom_enabled = self.scan_from_bottom_checkbox.isChecked()

        # Ensure unique_display_headers is updated before processing
        self.update_table_headers_and_content()

        for i, path in enumerate(file_paths):
            row_position = self.file_table_widget.rowCount()
            self.file_table_widget.insertRow(row_position)

            no_item = QTableWidgetItem(str(row_position + 1))
            no_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.file_table_widget.setItem(row_position, 0, no_item)

            file_name = os.path.basename(path)
            file_name_item = QTableWidgetItem(file_name)
            file_name_item.setData(Qt.ItemDataRole.UserRole, path)
            self.file_table_widget.setItem(row_position, 1, file_name_item)

            # Initialize columns for unique display headers
            for col_idx in range(len(self.unique_display_headers)):
                item = QTableWidgetItem("...")
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.file_table_widget.setItem(row_position, col_idx + 2, item)

            last_col_idx = 2 + len(self.unique_display_headers)
            item = QTableWidgetItem("...")
            item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
            self.file_table_widget.setItem(row_position, last_col_idx, item)

            worker = PdfProcessingWorker(
                path,
                self.list_teks_dicari, # Pass the full list of (search_text, display_text)
                current_dpi,
                process_from_bottom=process_from_bottom_enabled,
                stop_on_keyword_flag=process_from_bottom_enabled
            )
            worker.finished.connect(self.update_file_status_in_table)
            self.worker_threads.append(worker)
            worker.start()

    def update_file_status_in_table(self, pdf_path, results_dict, error_message, extracted_final_text):
        row_index = -1
        for r in range(self.file_table_widget.rowCount()):
            if self.file_table_widget.item(r, 1).data(Qt.ItemDataRole.UserRole) == pdf_path:
                row_index = r
                break

        if row_index != -1:
            if error_message:
                file_name_item = self.file_table_widget.item(row_index, 1)
                file_name_item.setForeground(QColor("#F78888"))

                # Mark all keyword columns as error
                for col_idx in range(len(self.unique_display_headers)):
                    detail_item = self.file_table_widget.item(row_index, col_idx + 2)
                    detail_item.setText("E")
                    detail_item.setForeground(QColor("#F78888"))

                last_col_idx = 2 + len(self.unique_display_headers)
                last_col_item = self.file_table_widget.item(row_index, last_col_idx)
                if not last_col_item: # Create if it doesn't exist
                    last_col_item = QTableWidgetItem()
                    self.file_table_widget.setItem(row_index, last_col_idx, last_col_item)

                last_col_item.setText(f"Error: {error_message[:50]}...")
                last_col_item.setToolTip(error_message)
                last_col_item.setForeground(QColor("#F78888"))
            else:
                # Update columns based on unique display headers and results_dict
                for col_idx, display_header in enumerate(self.unique_display_headers):
                    detail_item = self.file_table_widget.item(row_index, col_idx + 2)
                    # Pastikan results_dict memiliki kunci display_header
                    if display_header in results_dict and results_dict[display_header]: 
                        detail_item.setText("✓")
                        detail_item.setForeground(QColor("#00BFFF"))
                    else:
                        detail_item.setText("✗")
                        detail_item.setForeground(QColor("#F78888"))

                last_col_idx = 2 + len(self.unique_display_headers)
                last_col_item = self.file_table_widget.item(row_index, last_col_idx)
                if not last_col_item:
                    last_col_item = QTableWidgetItem()
                    self.file_table_widget.setItem(row_index, last_col_idx, last_col_item)

                if len(extracted_final_text) > 50:
                    display_text = extracted_final_text[:47] + "..."
                    last_col_item.setText(display_text)
                    last_col_item.setToolTip(extracted_final_text)
                else:
                    last_col_item.setText(extracted_final_text)
                last_col_item.setForeground(QColor("#C8CDD3"))
                last_col_item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

            self.worker_threads = [worker for worker in self.worker_threads if worker.pdf_path != pdf_path]

            if not self.worker_threads:
                self.status_label.setText("Semua file selesai diverifikasi.")
                self.save_button.setEnabled(True)

    def save_results_to_excel(self):
        if self.file_table_widget.rowCount() == 0:
            QMessageBox.warning(self, "Tidak Ada Data", "Tidak ada data untuk disimpan ke Excel.")
            return

        file_dialog = QFileDialog()
        excel_file_path, _ = file_dialog.getSaveFileName(
            self, "Simpan Hasil sebagai Excel", "hasil_verifikasi.xlsx", "Excel Files (*.xlsx)"
        )

        if excel_file_path:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Hasil Verifikasi PDF"

                headers = [self.file_table_widget.horizontalHeaderItem(col).text()
                           for col in range(self.file_table_widget.columnCount())]
                ws.append(headers)

                # --- Pengaturan Warna dan Style untuk Header Excel (Netral) ---
                header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                header_font = ExcelFont(bold=True, color="000000") # Font hitam
                thin_border = Border(left=Side(style='thin', color="BFBFBF"), # Border abu-abu terang
                                     right=Side(style='thin', color="BFBFBF"),
                                     top=Side(style='thin', color="BFBFBF"),
                                     bottom=Side(style='thin', color="BFBFBF"))

                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

                # Mendefinisikan warna font yang akan digunakan
                default_font = ExcelFont(color="000000") # Hitam
                green_font = ExcelFont(color="00B050") # Hijau untuk centang
                red_font = ExcelFont(color="FF0000")   # Merah untuk silang dan error

                for row_idx in range(self.file_table_widget.rowCount()):
                    row_data = []
                    for col_idx in range(self.file_table_widget.columnCount()):
                        item = self.file_table_widget.item(row_idx, col_idx)
                        if item:
                            cell_value = item.text()
                            row_data.append(cell_value)
                        else:
                            row_data.append("")
                    ws.append(row_data)

                    for col_idx, value in enumerate(row_data):
                        excel_cell = ws.cell(row=row_idx + 2, column=col_idx + 1)
                        excel_cell.font = default_font # Set font default (hitam)
                        excel_cell.alignment = Alignment(horizontal='center', vertical='center') # Default center alignment
                        excel_cell.border = thin_border # Menambahkan border tipis pada semua sel data

                        if col_idx == 1: # Kolom Nama File
                            excel_cell.alignment = Alignment(horizontal='left', vertical='center')
                            last_col_value = row_data[self.file_table_widget.columnCount() - 1]
                            if "Error:" in last_col_value:
                                excel_cell.font = red_font
                        elif 2 <= col_idx < (2 + len(self.unique_display_headers)): # Kolom kata kunci ('✓', '✗', 'E')
                            if value == "✓":
                                excel_cell.font = green_font
                            elif value == "✗" or value == "E": # Sekarang '✗' dan 'E' akan merah
                                excel_cell.font = red_font
                        elif col_idx == (2 + len(self.unique_display_headers)): # Kolom terakhir (EXTRACTION_KEYWORD)
                            excel_cell.alignment = Alignment(horizontal='left', vertical='center')
                            if "Error:" in value:
                                excel_cell.font = red_font
                        
                for col_idx in range(ws.max_column):
                    column_letter = ws.cell(row=1, column=col_idx + 1).column_letter
                    max_length = 0
                    # Ambil data dari tabel GUI untuk penyesuaian lebar kolom
                    if col_idx < self.file_table_widget.columnCount():
                        for row_in_gui in range(self.file_table_widget.rowCount()):
                            item = self.file_table_widget.item(row_in_gui, col_idx)
                            if item and item.text():
                                current_length = len(item.text())
                                if current_length > max_length:
                                    max_length = current_length
                        # Tambahkan lebar dari header juga
                        header_text = self.file_table_widget.horizontalHeaderItem(col_idx).text()
                        if len(header_text) > max_length:
                            max_length = len(header_text)
                    
                    adjusted_width = (max_length + 2)
                    if adjusted_width > 50: # Batasi lebar maksimum
                        adjusted_width = 50
                    elif adjusted_width < 10: # Batasi lebar minimum
                        adjusted_width = 10
                    ws.column_dimensions[column_letter].width = adjusted_width

                wb.save(excel_file_path)
                QMessageBox.information(self, "Berhasil", f"Hasil disimpan ke:\n{excel_file_path}")
            except Exception as e:
                QMessageBox.critical(self, "Error Menyimpan Excel", f"Gagal menyimpan file Excel:\n{e}")
        else:
            QMessageBox.information(self, "Penyimpanan Dibatalkan", "Penyimpanan file Excel dibatalkan.")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = PdfVerifierApp()
    window.show()
    sys.exit(app.exec())